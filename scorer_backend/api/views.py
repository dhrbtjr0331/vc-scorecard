from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.contrib.auth import authenticate, login
from django.contrib.auth import get_user_model
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import api_view, permission_classes
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.views import TokenRefreshView
from django.http import JsonResponse
from .models import Scorecard
import dropbox
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from django.conf import settings
import tempfile
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side


User = get_user_model()

def generate_jwt_tokens(user):
    """Utility function to generate JWT tokens for a user."""
    refresh = RefreshToken.for_user(user)
    return {
        'refresh': str(refresh),
        'access': str(refresh.access_token),
    }

def validate_registration_fields(username, password, retype_password, firstname, lastname):
    """Helper function to validate registration fields."""
    if not username:
        return "Username is required."
    if not password:
        return "Password is required."
    if not retype_password:
        return "Retype password is required."
    if not firstname:
        return "First name is required."
    if not lastname:
        return "Last name is required."
    return None

def validate_scorecard_fields(date, company_name, sector, investment_stage, alignment, team, market, product, potential_return, bold_excitement):
    """Helper function to validate registration fields."""
    if not date:
        return "Date is required."
    if not company_name:
        return "Company Name is required."
    if not sector:
        return "Sector is required."
    if not investment_stage:
        return "Investment Stage is required."
    if not alignment:
        return "Alignment is required."
    if not team:
        return "Team is required."
    if not market:
        return "Market is required."
    if not product:
        return "Product is required."
    if not potential_return:
        return "Potential_return is required."
    if not bold_excitement:
        return "Bold_excitement is required."
    return None

def getScore(alignment, team, market, product, potential_return, bold_excitement):
    score = (alignment + market + product + bold_excitement) * (team + potential_return) / 80.00
    return round(score, 2)

def fetch_excel_from_dropbox(dropbox_path, local_path):
    """Fetches an Excel file from Dropbox and saves it locally."""
    dbx = dropbox.Dropbox(settings.DROPBOX_ACCESS_TOKEN)
    print(f"Token being used: '{settings.DROPBOX_ACCESS_TOKEN}'")
    with open(local_path, "wb") as f:
        metadata, res = dbx.files_download(path=dropbox_path)
        f.write(res.content)
    print(f"File successfully fetched from Dropbox to: {local_path}")
    return local_path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

def append_and_format_to_excel(file_path, scorecard):
    """Appends a new scorecard row to the Excel file with grouped companies and their average score."""
    workbook = load_workbook(file_path)
    sheet = workbook.active

    name = f"{scorecard['scored_by']['first_name']} {scorecard['scored_by']['last_name']}"
    
    # Create the new row
    new_row = [
        name,
        scorecard["company_name"],
        scorecard["date"],
        scorecard["sector"],
        scorecard["investment_stage"],
        scorecard["scores"]["alignment"],
        scorecard["scores"]["team"],
        scorecard["scores"]["market"],
        scorecard["scores"]["product"],
        scorecard["scores"]["potential_return"],
        scorecard["scores"]["bold_excitement"],
        scorecard["score"]
    ]

    # Append the new row
    sheet.append(new_row)

    # Read all rows
    rows = list(sheet.iter_rows(values_only=True))
    header = rows[0]
    data = rows[1:]

    # Group rows by company name
    grouped_data = {}
    for row in data:
        company_name = row[1]  # Assuming Company Name is column B
        if company_name:
            grouped_data.setdefault(company_name, []).append(row)

    # Clear rows below the header
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)

    # Define a border style for the "Total Score" section
    border_style_1 = Border(
        left=Side(style="thick"),
        top=Side(style="thick"),
        bottom=Side(style="thick")
    )
    border_style_2 = Border(
        right=Side(style="thick"),
        top=Side(style="thick"),
        bottom=Side(style="thick")
    )

    # Write grouped data back with formatting
    for company, rows in grouped_data.items():
        scoreSum = 0
        count = 0
        if rows:  # Only process non-empty groups
            for row in rows:
                sheet.append(row)
                scoreSum += row[11]  # Assuming 'Score' column is at index 11
                count += 1

            # Calculate average score
            avg_score = round(scoreSum / count, 2) if count > 0 else 0

            # Get the row index for "Total Score"
            total_score_row = sheet.max_row + 1
            score_col = len(header)  # Last column (score column)
            label_col = score_col - 1  # Column before score

            # Merge cells for "Total Score"
            sheet.merge_cells(start_row=total_score_row, start_column=label_col, 
                              end_row=total_score_row, end_column=score_col - 1)

            # Set "Total Score" text
            total_score_cell = sheet.cell(row=total_score_row, column=label_col)
            total_score_cell.value = "Total Score"
            total_score_cell.alignment = Alignment(horizontal="center", vertical="center")
            total_score_cell.border = border_style_1  # Apply border

            # Set total score value in the last column
            score_cell = sheet.cell(row=total_score_row, column=score_col)
            score_cell.value = avg_score
            score_cell.alignment = Alignment(horizontal="center", vertical="center")
            score_cell.border = border_style_2  # Apply border

            # Append two blank rows after each company section
            sheet.append([""] * len(header))

    # Remove trailing blank rows
    while all(cell.value is None for cell in sheet[sheet.max_row]):
        sheet.delete_rows(sheet.max_row)

    # Format the header row
    for cell in sheet[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center")

    workbook.save(file_path)
    workbook.close()
    print(f"Excel file updated and saved locally at: {file_path}")

def upload_excel_to_dropbox(local_path, dropbox_path):
    """Uploads an Excel file to Dropbox."""
    dbx = dropbox.Dropbox(settings.DROPBOX_ACCESS_TOKEN)
    with open(local_path, "rb") as f:
        res = dbx.files_upload(f.read(), path=dropbox_path, mode=dropbox.files.WriteMode.overwrite)
    print(f"File successfully uploaded to Dropbox: {dropbox_path}")
    return f"Updated file uploaded to Dropbox at {dropbox_path}"

# Login View with JWT Token Generation
class LoginView(APIView):
    def post(self, request):
        username = request.data.get('username')
        password = request.data.get('password')
        
        user = authenticate(request, username=username, password=password)

        if user:
            # Generate JWT Tokens
            tokens = generate_jwt_tokens(user)
            return Response(tokens, status=status.HTTP_200_OK)

        return Response({'error': 'Invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)

# Register View
class RegisterView(APIView):
    def post(self, request):
        username = request.data.get('username')
        password = request.data.get('password')
        retype_password = request.data.get('retypePassword')
        firstname = request.data.get('firstname')
        lastname = request.data.get('lastname')

        # Validate fields
        validation_error = validate_registration_fields(username, password, retype_password, firstname, lastname)
        if validation_error:
            return Response({"error": validation_error}, status=status.HTTP_400_BAD_REQUEST)

        if password != retype_password:
            return Response({"error": "Passwords do not match."}, status=status.HTTP_400_BAD_REQUEST)

        if User.objects.filter(username=username).count() > 0:
            return Response({"error": "Username already exists."}, status=status.HTTP_409_CONFLICT)

        user = User.objects.create_user(username=username, password=password, first_name=firstname, last_name=lastname)

        tokens = generate_jwt_tokens(user)
        return Response({
            "message": "Registration successful.",
            **tokens,
        }, status=status.HTTP_201_CREATED)

# IsAuthenticated View to Check User Status
class IsAuthenticatedView(APIView):
    def get(self, request):
        if request.user.is_authenticated:
            return Response({
                "isAuthenticated": True,
                "user": {
                    "username": request.user.username,
                    "firstName": request.user.first_name,
                    "lastName": request.user.last_name
                }
            }, status=status.HTTP_200_OK)
        return Response({"isAuthenticated": False}, status=status.HTTP_401_UNAUTHORIZED)

# Logout View (Clear JWT on Frontend)
class LogoutView(APIView):
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        return Response({
            "message": "Logout successful. Please ensure that tokens are cleared on the client side to complete the logout process."
        }, status=status.HTTP_200_OK)
    
# Score Company View
class ScoreCompanyView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        date = request.data.get('date')
        company_name = request.data.get('company_name')
        sector = request.data.get('sector')
        investment_stage = request.data.get('investment_stage')
        alignment = int(request.data.get('alignment', 0))
        team = int(request.data.get('team', 0))
        market = int(request.data.get('market', 0))
        product = int(request.data.get('product', 0))
        potential_return = int(request.data.get('potential_return', 0))
        bold_excitement = int(request.data.get('bold_excitement', 0))

        # Validate data
        if not all([date, company_name, sector, investment_stage]):
            return Response({"error": "All fields are required"}, status=status.HTTP_400_BAD_REQUEST)

        # Calculate the overall score
        score = (alignment + market + product + bold_excitement) * (team + potential_return) / 80.0
        score = round(score, 2)

        # Create scorecard data
        scorecard = {
            "id": 1,  # Mock ID for debugging
            "date": date,
            "company_name": company_name,
            "sector": sector,
            "investment_stage": investment_stage,
            "scores": {
                "alignment": alignment,
                "team": team,
                "market": market,
                "product": product,
                "potential_return": potential_return,
                "bold_excitement": bold_excitement,
            },
            "score": score,
            "scored_by": {
                "first_name": request.user.first_name,
                "last_name": request.user.last_name,
            }
        }

        # Define paths
        dropbox_path = "/scorecards.xlsx"  # Ensure correct path
        local_path = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name

        try:
            # Step 1: Fetch file from Dropbox
            print(f"Fetching file from Dropbox: {dropbox_path}")
            fetch_excel_from_dropbox(dropbox_path, local_path)
            print(f"File successfully fetched from Dropbox: {local_path}")

            # Step 2: Update Excel file
            print("Appending and formatting the Excel file...")
            append_and_format_to_excel(local_path, scorecard)
            print(f"Excel file updated at: {local_path}")

            # Ensure file exists before upload
            if not os.path.exists(local_path):
                print(f"File does not exist at {local_path}, aborting upload.")
                return Response({"error": "Local file was not created properly"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            # Step 3: Upload file back to Dropbox
            print(f"Uploading updated file to Dropbox: {dropbox_path}")
            message = upload_excel_to_dropbox(local_path, dropbox_path)
            print(message)

        except Exception as e:
            print(f"An error occurred: {e}")
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        finally:
            # Clean up the local temporary file
            if os.path.exists(local_path):
                os.unlink(local_path)
                print(f"Temporary file cleaned up: {local_path}")

        return Response(scorecard, status=status.HTTP_201_CREATED)


class SectorListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        sectors = Scorecard.objects.values_list('sector', flat=True).distinct()
        return Response(list(sectors), status=status.HTTP_200_OK)

class CompanyNameListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        company_names = Scorecard.objects.values_list('company_name', flat=True).distinct()
        return Response(list(company_names), status=status.HTTP_200_OK)
    
class ScoredCompanyListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        scorecards = Scorecard.objects.select_related('scored_by').all()
        response_data = []
        for scorecard in scorecards:
            response_data.append({
                "id": scorecard.id,
                "date": scorecard.date,
                "company_name": scorecard.company_name,
                "sector": scorecard.sector,
                "investment_stage": scorecard.investment_stage,
                "alignment": scorecard.alignment,
                "team": scorecard.team,
                "market": scorecard.market,
                "product": scorecard.product,
                "potential_return": scorecard.potential_return,
                "bold_excitement": scorecard.bold_excitement,
                "score": scorecard.score,
                "scored_by": {
                    "first_name": scorecard.scored_by.first_name,
                    "last_name": scorecard.scored_by.last_name,
                } if scorecard.scored_by else None
            })
        return Response(response_data, status=status.HTTP_200_OK)
